"""
Accrual Entry Generator for GA Automation Pipeline
====================================================
Generates journal entries for accruals from three sources:
  Layer 1: Nexus pending invoices (AP-side accruals)
  Layer 2: Budget gap detection (accounts with budget but no GL activity)
  Layer 3: Historical pattern detection (recurring expenses from prior months)

Outputs:
  1. Yardi JE import file (Excel) â ready for direct upload
  2. Workpaper review schedule â DR/CR detail for review before posting

Each accrual generates a two-line entry:
  DR  [Expense GL Account]
  CR  211200 Accrued Expenses (standard accrual liability)
"""

import os
from datetime import datetime
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ââ Constants ââââââââââââââââââââââââââââââââââââââââââââââââ

AP_ACCRUAL_ACCOUNT = '211200'
AP_ACCRUAL_NAME = 'Accrued Expenses'

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
DOUBLE_BOTTOM = Border(bottom=Side(style='double'))

DARK_BLUE = '1F4E78'
MED_BLUE = '2E75B6'
LIGHT_BLUE = 'D6E4F0'
LIGHT_GRAY = 'F2F2F2'
WHITE = 'FFFFFF'


def _apply(cell, font=None, fill=None, fmt=None, border=None, align=None):
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    if border:
        cell.border = border
    if align:
        cell.alignment = align


def _hdr_font():
    return Font(name='Calibri', size=11, bold=True, color='FFFFFF')

def _hdr_fill():
    return PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type='solid')

def _subhdr_fill():
    return PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type='solid')


# ââ Layer 2: Budget gap detection ââââââââââââââââââââââââââââ

def detect_budget_gaps(gl_data, budget_data) -> List[Dict[str, Any]]:
    """
    Identify accounts that have a budget amount but zero GL activity,
    indicating a likely accrual candidate.

    Returns list of dicts: account_code, account_name, budget_amount, source='budget_gap'
    """
    candidates = []

    if not budget_data or not gl_data:
        return candidates

    # Build set of GL accounts with activity this period
    gl_active = set()
    if hasattr(gl_data, 'accounts'):
        for acct in gl_data.accounts:
            if abs(acct.net_change) > 0.01:
                gl_active.add(acct.account_code)

    # Check budget items
    budget_items = []
    if isinstance(budget_data, list):
        budget_items = budget_data
    elif hasattr(budget_data, 'line_items'):
        budget_items = budget_data.line_items

    for item in budget_items:
        if isinstance(item, dict):
            code = item.get('account_code', '')
            name = item.get('account_name', '')
            ptd_budget = item.get('ptd_budget', 0) or 0
            ptd_actual = item.get('ptd_actual', 0) or 0
        else:
            code = getattr(item, 'account_code', '')
            name = getattr(item, 'account_name', '')
            ptd_budget = getattr(item, 'ptd_budget', 0) or 0
            ptd_actual = getattr(item, 'ptd_actual', 0) or 0

        if not code or 'TOTAL' in str(name).upper():
            continue

        # Only expense accounts (5xxxxx-8xxxxx) with budget > $100 and no actual
        first_digit = code[0] if code else '0'
        if first_digit in ('5', '6', '7', '8') and abs(ptd_budget) > 100 and abs(ptd_actual) < 1:
            candidates.append({
                'account_code': code,
                'account_name': name,
                'budget_amount': abs(ptd_budget),
                'source': 'budget_gap',
                'description': f'Budget gap â {name} budgeted ${abs(ptd_budget):,.2f}, no GL activity',
            })

    return candidates


# ââ Layer 3: Historical pattern detection ââââââââââââââââââââ

def detect_historical_recurring(gl_data, budget_data) -> List[Dict[str, Any]]:
    """
    Identify recurring expense patterns by comparing GL beginning balance
    (YTD proxy) to budget. If an account had YTD activity through the prior
    month but nothing this month, it may need an accrual.

    Uses beginning_balance as a proxy for prior-month YTD activity.
    If beginning_balance shows consistent prior activity but net_change is
    zero, flag as a recurring accrual candidate.

    Returns list of dicts: account_code, account_name, estimated_amount, source='historical'
    """
    candidates = []

    if not gl_data or not hasattr(gl_data, 'accounts'):
        return candidates

    # Determine current month number from period
    period_str = getattr(gl_data.metadata, 'period', '') if hasattr(gl_data, 'metadata') else ''
    month_num = 1
    if '-' in period_str:
        month_map = {
            'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
            'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12,
        }
        month_name = period_str.split('-')[0]
        month_num = month_map.get(month_name, 1)

    for acct in gl_data.accounts:
        code = acct.account_code
        first_digit = code[0] if code else '0'

        # Only expense accounts
        if first_digit not in ('5', '6', '7', '8'):
            continue

        # Skip if there's activity this month
        if abs(acct.net_change) > 0.01:
            continue

        # Check if beginning balance suggests recurring prior activity
        # Beginning balance for expense accounts is YTD through prior month
        begin = abs(acct.beginning_balance)
        if begin < 100 or month_num <= 1:
            continue

        # Estimate monthly amount from YTD / months elapsed
        prior_months = month_num - 1
        est_monthly = begin / prior_months

        # Only flag if estimated monthly > $500 (material recurring expense)
        if est_monthly >= 500:
            candidates.append({
                'account_code': code,
                'account_name': acct.account_name,
                'estimated_amount': round(est_monthly, 2),
                'ytd_prior': begin,
                'months_prior': prior_months,
                'source': 'historical',
                'description': f'Recurring â {acct.account_name} avg ${est_monthly:,.0f}/mo ({prior_months} prior months), no activity this period',
            })

    return candidates


# ââ Build JE lines from all sources âââââââââââââââââââââââââ

def build_accrual_entries(nexus_data: list, period: str = '',
                          property_name: str = '',
                          status_filter: list = None,
                          gl_data=None, budget_data=None) -> List[Dict[str, Any]]:
    """
    Build accrual journal entry lines from three sources:
      Layer 1: Nexus pending invoices
      Layer 2: Budget gap detection (gl_data + budget_data required)
      Layer 3: Historical recurring detection (gl_data required)

    Args:
        nexus_data: List of invoice dicts from Nexus parser
        period: Accounting period (e.g., 'Feb-2026')
        property_name: Property name for the JE header
        status_filter: List of invoice statuses to include.
                       Default: include all invoices (pending + approved).

    Returns:
        List of JE line dicts with keys:
          je_number, line, date, account_code, account_name,
          description, reference, debit, credit, vendor, invoice_number
    """
    invoices = nexus_data if isinstance(nexus_data, list) else []

    if status_filter:
        invoices = [inv for inv in invoices
                    if (inv.get('invoice_status', '') or '').lower()
                    in [s.lower() for s in status_filter]]

    je_lines = []
    je_num = 1

    for inv in invoices:
        vendor = inv.get('vendor', '')
        inv_num = inv.get('invoice_number', '')
        inv_date = inv.get('invoice_date', '')
        gl_account = inv.get('gl_account', '')
        gl_category = inv.get('gl_category', '')
        description = inv.get('line_description', '')
        amount = inv.get('amount', 0) or 0

        if amount == 0:
            continue

        # Format date
        if isinstance(inv_date, datetime):
            date_str = inv_date.strftime('%m/%d/%Y')
        elif isinstance(inv_date, str):
            date_str = inv_date
        else:
            date_str = ''

        # Build description for JE
        je_desc = f"Accrual â {vendor}"
        if inv_num:
            je_desc += f" #{inv_num}"
        if description:
            je_desc += f" â {description[:50]}"

        je_id = f"ACC-{je_num:04d}"

        # DR line: Expense account
        je_lines.append({
            'je_number': je_id,
            'line': 1,
            'date': date_str,
            'account_code': gl_account,
            'account_name': gl_category or description[:30],
            'description': je_desc,
            'reference': inv_num,
            'debit': abs(amount),
            'credit': 0,
            'vendor': vendor,
            'invoice_number': inv_num,
        })

        # CR line: AP Accrual (211200)
        je_lines.append({
            'je_number': je_id,
            'line': 2,
            'date': date_str,
            'account_code': AP_ACCRUAL_ACCOUNT,
            'account_name': AP_ACCRUAL_NAME,
            'description': je_desc,
            'reference': inv_num,
            'debit': 0,
            'credit': abs(amount),
            'vendor': vendor,
            'invoice_number': inv_num,
        })

        je_num += 1

    # ââ Layer 2: Budget gap accruals ââ
    if gl_data and budget_data:
        budget_gaps = detect_budget_gaps(gl_data, budget_data)
        # Don't duplicate accounts already covered by Nexus
        nexus_accounts = set()
        for line in je_lines:
            if line['line'] == 1:  # DR lines only
                nexus_accounts.add(line['account_code'])

        for gap in budget_gaps:
            if gap['account_code'] in nexus_accounts:
                continue

            je_id = f"BGA-{je_num:04d}"
            je_desc = f"Budget gap accrual â {gap['account_name']}"

            je_lines.append({
                'je_number': je_id,
                'line': 1,
                'date': '',
                'account_code': gap['account_code'],
                'account_name': gap['account_name'],
                'description': je_desc,
                'reference': 'BUDGET-GAP',
                'debit': gap['budget_amount'],
                'credit': 0,
                'vendor': '[Budget Gap]',
                'invoice_number': '',
                'source': 'budget_gap',
            })
            je_lines.append({
                'je_number': je_id,
                'line': 2,
                'date': '',
                'account_code': AP_ACCRUAL_ACCOUNT,
                'account_name': AP_ACCRUAL_NAME,
                'description': je_desc,
                'reference': 'BUDGET-GAP',
                'debit': 0,
                'credit': gap['budget_amount'],
                'vendor': '[Budget Gap]',
                'invoice_number': '',
                'source': 'budget_gap',
            })
            je_num += 1

    # ââ Layer 3: Historical recurring accruals ââ
    if gl_data:
        historicals = detect_historical_recurring(gl_data, budget_data)
        covered_accounts = set()
        for line in je_lines:
            if line['line'] == 1:
                covered_accounts.add(line['account_code'])

        for hist in historicals:
            if hist['account_code'] in covered_accounts:
                continue

            je_id = f"REC-{je_num:04d}"
            je_desc = f"Recurring accrual â {hist['account_name']} (est. ${hist['estimated_amount']:,.0f}/mo)"

            je_lines.append({
                'je_number': je_id,
                'line': 1,
                'date': '',
                'account_code': hist['account_code'],
                'account_name': hist['account_name'],
                'description': je_desc,
                'reference': 'RECURRING',
                'debit': hist['estimated_amount'],
                'credit': 0,
                'vendor': '[Historical Recurring]',
                'invoice_number': '',
                'source': 'historical',
            })
            je_lines.append({
                'je_number': je_id,
                'line': 2,
                'date': '',
                'account_code': AP_ACCRUAL_ACCOUNT,
                'account_name': AP_ACCRUAL_NAME,
                'description': je_desc,
                'reference': 'RECURRING',
                'debit': 0,
                'credit': hist['estimated_amount'],
                'vendor': '[Historical Recurring]',
                'invoice_number': '',
                'source': 'historical',
            })
            je_num += 1

    return je_lines


# ââ Generate Yardi JE import file ââââââââââââââââââââââââââââ

def generate_yardi_je_import(je_lines: List[Dict], output_path: str,
                              period: str = '', property_name: str = '') -> str:
    """
    Generate a Yardi-compatible journal entry import file (Excel).

    Yardi JE import expects columns:
      Property, Journal #, Date, Account, Description, Reference, Debit, Credit

    Args:
        je_lines: List of JE line dicts from build_accrual_entries()
        output_path: Where to write the Excel file
        period: Accounting period
        property_name: Property code/name

    Returns:
        Output path
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Journal Entries'

    # Header row
    headers = ['Property', 'Journal #', 'Date', 'Account', 'Description',
               'Reference', 'Debit', 'Credit']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        _apply(c, font=_hdr_font(), fill=_hdr_fill(), border=THIN_BORDER,
               align=Alignment(horizontal='center', vertical='center'))

    # Data rows
    prop_code = property_name.split()[0] if property_name else 'REVLABS'

    for ri, line in enumerate(je_lines, 2):
        alt_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type='solid') if ri % 2 == 0 else None

        ws.cell(row=ri, column=1, value=prop_code)
        ws.cell(row=ri, column=2, value=line['je_number'])
        ws.cell(row=ri, column=3, value=line['date'])
        ws.cell(row=ri, column=4, value=line['account_code'])
        ws.cell(row=ri, column=5, value=line['description'])
        ws.cell(row=ri, column=6, value=line['reference'])

        c_dr = ws.cell(row=ri, column=7, value=line['debit'])
        c_dr.number_format = '$#,##0.00'
        c_cr = ws.cell(row=ri, column=8, value=line['credit'])
        c_cr.number_format = '$#,##0.00'

        for ci in range(1, 9):
            ws.cell(row=ri, column=ci).border = THIN_BORDER
            if alt_fill:
                ws.cell(row=ri, column=ci).fill = alt_fill

    # Totals row
    total_row = len(je_lines) + 2
    ws.cell(row=total_row, column=6, value='TOTAL').font = Font(name='Calibri', size=11, bold=True)
    total_dr = sum(l['debit'] for l in je_lines)
    total_cr = sum(l['credit'] for l in je_lines)
    c_dr = ws.cell(row=total_row, column=7, value=total_dr)
    c_dr.number_format = '$#,##0.00'
    c_dr.font = Font(name='Calibri', size=11, bold=True)
    c_dr.border = DOUBLE_BOTTOM
    c_cr = ws.cell(row=total_row, column=8, value=total_cr)
    c_cr.number_format = '$#,##0.00'
    c_cr.font = Font(name='Calibri', size=11, bold=True)
    c_cr.border = DOUBLE_BOTTOM

    # Validation check
    balance_row = total_row + 1
    ws.cell(row=balance_row, column=6, value='Balance Check').font = Font(name='Calibri', size=10, italic=True)
    diff = total_dr - total_cr
    c_bal = ws.cell(row=balance_row, column=7, value=diff)
    c_bal.number_format = '$#,##0.00'
    c_bal.font = Font(name='Calibri', size=10, italic=True,
                      color='008000' if abs(diff) < 0.01 else 'FF0000')

    # Auto column widths
    for col in range(1, 9):
        letter = chr(64 + col)
        best = 12
        for cell in ws[letter]:
            try:
                if cell.value:
                    best = max(best, len(str(cell.value)) + 2)
            except:
                pass
        ws.column_dimensions[letter].width = min(best, 45)

    wb.save(output_path)
    return output_path


# ââ Add review tab to workpapers âââââââââââââââââââââââââââââ

def write_accrual_entries_workpaper_tab(wb: Workbook, je_lines: List[Dict],
                                         period: str = '', property_name: str = ''):
    """
    Add an 'Accrual Entries' review tab to an existing workbook.
    Shows JE detail with DR/CR, grouped by vendor, for review before posting.

    Args:
        wb: Existing workbook to add the tab to
        je_lines: List of JE line dicts from build_accrual_entries()
        period: Accounting period
        property_name: Property name
    """
    ws = wb.create_sheet('Accrual Entries')

    # Title
    row = 1
    c = ws.cell(row=row, column=1, value=f'Accrual Journal Entries â {property_name}')
    c.font = Font(name='Calibri', size=14, bold=True, color=DARK_BLUE)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    row += 1

    c = ws.cell(row=row, column=1,
                value=f'Period: {period}  |  CR Account: {AP_ACCRUAL_ACCOUNT} {AP_ACCRUAL_NAME}  |  Prepared: {datetime.now().strftime("%m/%d/%Y")}')
    c.font = Font(name='Calibri', size=11, italic=True, color='666666')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    row += 1

    # Summary
    total_entries = len([l for l in je_lines if l['line'] == 1])
    total_amount = sum(l['debit'] for l in je_lines)
    c = ws.cell(row=row, column=1,
                value=f'Total Entries: {total_entries}  |  Total Amount: ${total_amount:,.2f}')
    c.font = Font(name='Calibri', size=11, bold=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    row += 2

    # Column headers
    headers = ['JE #', 'Line', 'Vendor', 'Invoice #', 'Date',
               'Account', 'Description', 'Debit', 'Credit']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        _apply(c, font=_hdr_font(), fill=_hdr_fill(), border=THIN_BORDER,
               align=Alignment(horizontal='center', vertical='center', wrap_text=True))
    row += 1

    # Data rows
    current_je = None
    for i, line in enumerate(je_lines):
        alt = (i // 2) % 2 == 1  # Alternate every JE pair
        fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type='solid') if alt else None

        # JE group separator
        if line['je_number'] != current_je:
            current_je = line['je_number']

        ws.cell(row=row, column=1, value=line['je_number'])
        ws.cell(row=row, column=2, value=line['line'])
        ws.cell(row=row, column=3, value=line['vendor'] if line['line'] == 1 else '')
        ws.cell(row=row, column=4, value=line['invoice_number'] if line['line'] == 1 else '')
        ws.cell(row=row, column=5, value=line['date'] if line['line'] == 1 else '')
        ws.cell(row=row, column=6, value=line['account_code'])

        # Shorten description for CR line
        desc = line['description'] if line['line'] == 1 else f"  CR {AP_ACCRUAL_ACCOUNT}"
        ws.cell(row=row, column=7, value=desc)

        c_dr = ws.cell(row=row, column=8, value=line['debit'] if line['debit'] > 0 else '')
        if line['debit'] > 0:
            c_dr.number_format = '$#,##0.00'

        c_cr = ws.cell(row=row, column=9, value=line['credit'] if line['credit'] > 0 else '')
        if line['credit'] > 0:
            c_cr.number_format = '$#,##0.00'

        for ci in range(1, 10):
            ws.cell(row=row, column=ci).border = THIN_BORDER
            if fill:
                ws.cell(row=row, column=ci).fill = fill

        row += 1

    # Totals
    row += 1
    ws.cell(row=row, column=7, value='TOTAL').font = Font(name='Calibri', size=11, bold=True)
    total_dr = sum(l['debit'] for l in je_lines)
    total_cr = sum(l['credit'] for l in je_lines)
    c_dr = ws.cell(row=row, column=8, value=total_dr)
    c_dr.number_format = '$#,##0.00'
    c_dr.font = Font(name='Calibri', size=11, bold=True)
    c_dr.border = DOUBLE_BOTTOM
    c_cr = ws.cell(row=row, column=9, value=total_cr)
    c_cr.number_format = '$#,##0.00'
    c_cr.font = Font(name='Calibri', size=11, bold=True)
    c_cr.border = DOUBLE_BOTTOM

    # Balance check
    row += 1
    diff = total_dr - total_cr
    ws.cell(row=row, column=7, value='Balance Check').font = Font(name='Calibri', size=10, italic=True)
    c = ws.cell(row=row, column=8, value=diff)
    c.number_format = '$#,##0.00'
    c.font = Font(name='Calibri', size=10, bold=True,
                  color='008000' if abs(diff) < 0.01 else 'FF0000')

    # Account summary section
    row += 3
    c = ws.cell(row=row, column=1, value='Account Summary')
    c.font = Font(name='Calibri', size=12, bold=True, color=DARK_BLUE)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1

    sum_headers = ['Account Code', 'Description', 'Total Debit', 'Entry Count']
    for ci, h in enumerate(sum_headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        _apply(c, font=Font(name='Calibri', size=11, bold=True, color='000000'),
               fill=_subhdr_fill(), border=THIN_BORDER)
    row += 1

    # Aggregate by GL account (DR side only)
    acct_totals = {}
    for line in je_lines:
        if line['debit'] > 0:
            code = line['account_code']
            if code not in acct_totals:
                acct_totals[code] = {'name': line['account_name'], 'total': 0, 'count': 0}
            acct_totals[code]['total'] += line['debit']
            acct_totals[code]['count'] += 1

    for code, data in sorted(acct_totals.items()):
        ws.cell(row=row, column=1, value=code); ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=data['name']); ws.cell(row=row, column=2).border = THIN_BORDER
        c = ws.cell(row=row, column=3, value=data['total'])
        c.number_format = '$#,##0.00'
        c.border = THIN_BORDER
        ws.cell(row=row, column=4, value=data['count']); ws.cell(row=row, column=4).border = THIN_BORDER
        row += 1

    # Auto-width
    for col in range(1, 10):
        letter = chr(64 + col) if col <= 26 else 'A'
        best = 12
        for cell in ws[letter]:
            try:
                if cell.value:
                    best = max(best, len(str(cell.value)) + 2)
            except:
                pass
        ws.column_dimensions[letter].width = min(best, 50)

    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['G'].width = 45
    ws.sheet_properties.tabColor = '7030A0'  # Purple for accrual entries
