"""
Parser for Kardin annual budget files (.xlsx format).

Expected file format (qryExportData sheet):
- Row 1 (headers): PropID, PropName, PCID, ProfitCenterName, DataType, StartDate,
                   PrimaryAccountTypeID, PrimaryAccountType, ChartID, ChartName,
                   SubChart, Code, AllocationName, Description, Journal,
                   M1, M2, M3, M4, M5, M6, M7, M8, M9, M10, M11, M12, MTotal,
                   StartYear, LedgerType, ReportSort
- Row 2+: Budget data rows with monthly values (M1-M12) and total (MTotal)

The parser handles:
- Extracting monthly budget values for each account/allocation
- Converting month columns (M1-M12, MTotal) to numeric values
- Mapping account codes to descriptions
- Handling different allocation names (profit centers)
"""

import openpyxl
from typing import List, Dict, Tuple, Any
from datetime import datetime


def parse(filepath: str) -> List[Dict[str, Any]]:
    """
    Parse a Kardin annual budget file and return list of budget records.

    Args:
        filepath: Path to .xlsx file

    Returns:
        List of dictionaries with keys:
        - prop_id: Property ID
        - prop_name: Property name
        - profit_center: Profit center name
        - account_code: GL Chart ID/Code
        - account_name: GL Chart name
        - account_type: Primary account type
        - allocation_name: Allocation/sub-budget name
        - description: Budget line description
        - subchart: Subchart identifier
        - m1 through m12: Monthly budget values (float)
        - m_total: Total annual budget (float)
    """
    workbook = openpyxl.load_workbook(filepath)

    # Find the data sheet (typically 'qryExportData')
    worksheet = None
    for sheet_name in workbook.sheetnames:
        if 'qryExportData' in sheet_name or 'data' in sheet_name.lower():
            worksheet = workbook[sheet_name]
            break

    if worksheet is None:
        worksheet = workbook.active

    records = []

    # Get header row (row 1)
    header_row = list(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    header_map = {cell: idx for idx, cell in enumerate(header_row)}

    # Parse data rows (starting from row 2)
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), 2):
        # Skip completely empty rows
        if all(cell is None or cell == '' for cell in row):
            continue

        try:
            # Extract key fields
            prop_id = row[header_map.get('PropID', 0)] or ''
            prop_name = row[header_map.get('PropName', 1)] or ''
            profit_center = row[header_map.get('ProfitCenterName', 3)] or ''
            account_code = row[header_map.get('ChartID', 8)] or ''
            account_name = row[header_map.get('ChartName', 9)] or ''
            account_type = row[header_map.get('PrimaryAccountType', 7)] or ''
            allocation_name = row[header_map.get('AllocationName', 12)] or ''
            description = row[header_map.get('Description', 13)] or ''
            subchart = row[header_map.get('SubChart', 10)] or ''

            # Extract monthly values
            months = {}
            for month_num in range(1, 13):
                col_name = f'M{month_num}'
                col_idx = header_map.get(col_name)
                if col_idx is not None:
                    months[col_name] = _parse_amount(row[col_idx])
                else:
                    months[col_name] = 0.0

            # Extract total
            m_total = _parse_amount(row[header_map.get('MTotal')])

            record = {
                'prop_id': str(prop_id),
                'prop_name': str(prop_name),
                'profit_center': str(profit_center) if profit_center else None,
                'account_code': str(account_code),
                'account_name': str(account_name),
                'account_type': str(account_type) if account_type else None,
                'allocation_name': str(allocation_name),
                'description': str(description),
                'subchart': str(subchart) if subchart else None,
                **months,
                'm_total': m_total,
            }

            records.append(record)

        except Exception:
            # Skip rows with parsing errors
            continue

    return records


def validate(filepath: str) -> Tuple[bool, List[str]]:
    """
    Validate that a file has the expected Kardin budget format.

    Args:
        filepath: Path to .xlsx file

    Returns:
        Tuple of (is_valid, list_of_issues)
    """
    issues = []

    try:
        workbook = openpyxl.load_workbook(filepath)
    except Exception as e:
        return False, [f"Cannot open file: {str(e)}"]

    # Check for expected sheet
    found_data_sheet = False
    for sheet_name in workbook.sheetnames:
        if 'qryExportData' in sheet_name or 'data' in sheet_name.lower():
            found_data_sheet = True
            worksheet = workbook[sheet_name]
            break

    if not found_data_sheet:
        worksheet = workbook.active
        if worksheet is None:
            return False, ["No sheets found in workbook"]

    # Check header row
    header_row = list(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    header_str = ' '.join(str(h) for h in header_row if h)

    required_columns = ['PropID', 'PropName', 'ChartID', 'Description']
    for col_name in required_columns:
        if col_name not in header_str:
            issues.append(f"Missing expected column: {col_name}")

    # Check for monthly columns
    monthly_cols = [f'M{i}' for i in range(1, 13)]
    for col_name in monthly_cols:
        if col_name not in header_str:
            issues.append(f"Missing expected monthly column: {col_name}")

    # Check for data rows
    data_count = 0
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=100, values_only=True), 2):
        if any(cell is not None and cell != '' for cell in row):
            data_count += 1

    if data_count == 0:
        issues.append("No data rows found (empty worksheet)")

    return len(issues) == 0, issues


def _parse_amount(value: Any) -> float:
    """
    Parse amount value from Excel cell.

    Handles numbers, strings, and None.

    Returns float or 0.0 if cannot parse
    """
    if value is None or value == '':
        return 0.0

    if isinstance(value, (int, float)):
        return float(value)

    if isinstance(value, str):
        value = value.strip()
        if not value:
            return 0.0
        try:
            return float(value)
        except ValueError:
            return 0.0

    return 0.0


if __name__ == '__main__':
    import sys

    if len(sys.argv) > 1:
        filepath = sys.argv[1]

        # Validate
        is_valid, issues = validate(filepath)
        print(f"Validation: {'PASS' if is_valid else 'FAIL'}")
        if issues:
            for issue in issues:
                print(f"  - {issue}")

        # Parse
        records = parse(filepath)
        print(f"\nTotal budget records parsed: {len(records)}")

        if records:
            # Group by account and show summary
            by_account = {}
            for rec in records:
                account = rec['account_code']
                if account not in by_account:
                    by_account[account] = {'name': rec['account_name'], 'total': 0.0}
                by_account[account]['total'] += rec['m_total']

            total_budget = sum(by_account[a]['total'] for a in by_account)
            print(f"Total annual budget: ${total_budget:,.2f}")

            print("\nTop 10 accounts by budget:")
            sorted_accounts = sorted(by_account.items(), key=lambda x: x[1]['total'], reverse=True)
            for account, data in sorted_accounts[:10]:
                print(f"  {account} ({data['name'][:50]}): ${data['total']:,.2f}")
        else:
            print("No budget records found")
    else:
        print("Usage: python kardin_budget.py <filepath>")
