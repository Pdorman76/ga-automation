"""
Yardi Rent Roll Parser

Parses Yardi Rent Roll (Tenancy Schedule) export files with unit-level lease data:
- Row 1: Report title "Tenancy Schedule II"
- Row 2: Property, AsOf date, and notes
- Row 3+: Column headers (spread across multiple rows for multi-level headers)
- Row 7+: Lease and unit data (may have continuation rows for rent steps)

Expected main columns:
  Property, Building, Floor, Unit Code, Unit Type, Unit Area, Lease, Customer,
  Lease From, Lease To, Term, Tenancy, Lease Area, Annual Rent, Annual Rent/Area,
  Lease Type, LOC Amount, Rent, Start Date, Unit, Area Label, Area, Rent Step,
  Monthly, Rent Step, Annual, Management Fee, Annual Gross, Recov. Type, Base Yr,
  Base Amt

Features:
- Handles multi-row header structure
- Parses unit/lease information
- Handles rent step continuation rows
- Extracts dates and converts to ISO format
- Normalizes numeric values
- Groups rent steps by lease unit
"""

from openpyxl import load_workbook
from datetime import datetime
from typing import List, Dict, Tuple, Optional


def parse(filepath: str) -> List[Dict]:
    """
    Parse a Yardi Rent Roll export file.

    Args:
        filepath: Path to the Excel file

    Returns:
        List of dictionaries representing lease/unit records

    Raises:
        FileNotFoundError: If file does not exist
        ValueError: If file structure is invalid
    """
    try:
        wb = load_workbook(filepath)
    except Exception as e:
        raise FileNotFoundError(f"Cannot open file: {filepath}") from e

    ws = wb.active
    data = []

    # Meta information
    metadata = _extract_metadata(ws)

    # Extract headers from row 3 (main headers)
    headers = _extract_headers(ws, row=3)

    if not headers:
        raise ValueError("Cannot extract headers from Rent Roll file")

    # Detect the property name from row 6 or metadata to identify primary rows.
    # The property name prefix in column 1 distinguishes unit rows from
    # continuation rows (rent steps). We detect it dynamically rather than
    # hard-coding a specific property name.
    property_prefix = _detect_property_prefix(ws, metadata)

    # Process data rows starting from row 7
    # Note: row 6 contains property header, so actual unit data starts at 7
    current_unit = None

    for row_num in range(7, ws.max_row + 1):
        row = ws[row_num]
        row_values = [cell.value for cell in row]

        # Skip completely empty rows
        if all(v is None for v in row_values):
            continue

        # Check if this is a new unit record (column 1 has Property info)
        # Primary unit rows have a non-None value in column 1 that matches
        # the property prefix, or if no prefix detected, any non-None col 1
        # with data in multiple columns (not just a section header)
        is_primary = False
        if row_values[0] is not None:
            col1 = str(row_values[0]).strip()
            if property_prefix and col1.startswith(property_prefix):
                is_primary = True
            elif not property_prefix:
                # Fallback: primary rows have data in multiple columns
                non_none = sum(1 for v in row_values[:10] if v is not None)
                is_primary = non_none >= 3

        if is_primary:
            # This is a primary unit row
            record = _build_unit_record(headers, row_values, metadata)
            if record:
                data.append(record)
                current_unit = record
        elif row_values[0] is None and any(v is not None for v in row_values):
            # This is a continuation row (rent step)
            # Try to extract rent step data and append to current unit
            if current_unit is not None:
                # Extract rent step info from this row
                rent_step_data = _extract_rent_step_data(headers, row_values)
                if rent_step_data:
                    # Add as separate rent step record linked to the lease
                    step_record = dict(current_unit)
                    step_record.update(rent_step_data)
                    data.append(step_record)

    return data


def validate(filepath: str) -> Tuple[bool, List[str]]:
    """
    Validate that a file has the expected Rent Roll structure.

    Args:
        filepath: Path to the Excel file

    Returns:
        Tuple of (is_valid: bool, issues: list of error strings)
    """
    issues = []

    try:
        wb = load_workbook(filepath)
    except Exception as e:
        return False, [f"Cannot open file: {e}"]

    ws = wb.active

    # Check for expected title row
    if not ws.cell(1, 1).value or "Tenancy Schedule" not in str(ws.cell(1, 1).value):
        issues.append("Row 1 missing 'Tenancy Schedule II' title")

    # Check for property header in row 2
    if not ws.cell(2, 1).value or "Property" not in str(ws.cell(2, 1).value):
        issues.append("Row 2 missing property information")

    # Check headers
    headers = _extract_headers(ws, row=3)
    if not headers:
        issues.append("Cannot extract headers from row 3")
    else:
        # Should have at least Property, Unit Code, Customer
        if len(headers) < 3:
            issues.append("Expected at least 3 header columns")

    return len(issues) == 0, issues


def _detect_property_prefix(ws, metadata: Dict) -> Optional[str]:
    """Detect the property name prefix used in column 1 of data rows.

    Strategy:
    1. Check metadata for property name (from row 2 header).
    2. Scan rows 7-15 for the first non-None value in column 1 that
       appears to be a property/building name (not a section header).
    3. Use the common prefix of column 1 values as the property prefix.
    """
    # Try metadata first
    prop = metadata.get('property', '')
    if prop:
        # The property code is typically short (e.g., "REVLABS").
        # Look at actual data to find the full display name.
        pass

    # Scan first few data rows to find the property prefix
    col1_values = []
    for row_num in range(7, min(20, ws.max_row + 1)):
        val = ws.cell(row_num, 1).value
        if val is not None:
            col1_values.append(str(val).strip())

    if not col1_values:
        return None

    # If all values share a common prefix, use that
    if len(col1_values) >= 2:
        prefix = col1_values[0]
        for val in col1_values[1:]:
            # Find common prefix
            common = ''
            for a, b in zip(prefix, val):
                if a == b:
                    common += a
                else:
                    break
            prefix = common
        # Trim to last space to get a clean word boundary
        if ' ' in prefix:
            prefix = prefix[:prefix.rindex(' ')].strip()
        if len(prefix) >= 3:
            return prefix

    # Fallback: use the first value as-is (single row or no common prefix)
    return col1_values[0] if col1_values else None


def _extract_metadata(ws) -> Dict:
    """Extract metadata from title rows (rows 1-2)."""
    metadata = {}

    # Row 1: Title
    title = ws.cell(1, 1).value
    if title:
        metadata['report_title'] = str(title).strip()

    # Row 2: Property and AsOf date
    prop_line = ws.cell(2, 1).value
    if prop_line:
        prop_str = str(prop_line)
        metadata['report_details'] = prop_str

        # Try to extract property code and date
        if "Property:" in prop_str:
            parts = prop_str.split("As of Date:")
            if len(parts) > 1:
                prop_part = parts[0].split("Property:")[1].strip()
                metadata['property'] = prop_part.split()[0] if prop_part else None

                date_part = parts[1].strip().split()[0] if len(parts) > 1 else None
                if date_part:
                    metadata['as_of_date'] = date_part

    return metadata


def _extract_headers(ws, row: int) -> List[str]:
    """Extract and clean headers from a specific row."""
    headers = []
    for cell in ws[row]:
        value = cell.value
        if value:
            # Clean up whitespace and newlines
            clean_value = str(value).replace('\n', ' ').strip()
            headers.append(clean_value)
        else:
            headers.append(None)
    return headers


def _build_unit_record(headers: List[str], row_values: List, metadata: Dict) -> Optional[Dict]:
    """Build a complete unit record from header and value rows."""
    if not row_values or all(v is None for v in row_values):
        return None

    record = {}

    # Build the record with available values
    for i, header in enumerate(headers):
        if header and i < len(row_values):
            value = row_values[i]
            # Use normalized key names
            key = _normalize_header_name(header)
            record[key] = _normalize_value(value)

    # Add metadata
    record.update(metadata)

    return record if any(v is not None for v in record.values()) else None


def _extract_rent_step_data(headers: List[str], row_values: List) -> Optional[Dict]:
    """Extract rent step information from a continuation row."""
    rent_step_data = {}

    # Look for rent step columns in the row
    for i, header in enumerate(headers):
        if header and i < len(row_values):
            value = row_values[i]
            if value is not None:
                key = _normalize_header_name(header)
                rent_step_data[key] = _normalize_value(value)

    return rent_step_data if rent_step_data else None


def _normalize_header_name(header: str) -> str:
    """Convert header name to normalized snake_case key."""
    if not header:
        return ""

    # Remove special characters and convert to lowercase with underscores
    header = header.strip().lower()
    header = header.replace(' ', '_').replace('/', '_').replace('-', '_')
    # Remove duplicate underscores
    while '__' in header:
        header = header.replace('__', '_')
    return header.strip('_')


def _normalize_value(value):
    """Normalize values for consistent output."""
    if value is None:
        return None

    # Convert datetime to ISO format string
    if isinstance(value, datetime):
        return value.isoformat()

    # Handle numeric values
    if isinstance(value, (int, float)):
        return value

    # Handle strings - strip whitespace
    if isinstance(value, str):
        return value.strip()

    return value


def _normalize_numeric(value):
    """Normalize numeric values, handling None and strings."""
    if value is None:
        return None

    if isinstance(value, (int, float)):
        return value

    if isinstance(value, str):
        try:
            if '.' not in value:
                return int(value)
            return float(value)
        except (ValueError, AttributeError):
            return None

    return value


if __name__ == "__main__":
    import sys
    import json

    if len(sys.argv) < 2:
        print("Usage: python yardi_rent_roll.py <filepath>")
        sys.exit(1)

    filepath = sys.argv[1]

    # Validate
    is_valid, issues = validate(filepath)
    if not is_valid:
        print(f"Validation errors:")
        for issue in issues:
            print(f"  - {issue}")
        sys.exit(1)

    # Parse
    data = parse(filepath)
    print(f"Successfully parsed {len(data)} rent roll records")
    print(f"\nSample records (first 2 entries):")
    for i, record in enumerate(data[:2]):
        print(f"\nRecord {i+1}:")
        # Print a subset of keys to avoid overwhelming output
        sample = {k: v for k, v in list(record.items())[:10]}
        print(json.dumps(sample, indent=2, default=str))
